# -*- coding: utf-8 -*-
"""Excel 导出模块 - 用 openpyxl 生成失物清单 Excel 文件。"""

import io
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# 表头：(数据库字段, 中文表头)
COLUMNS = [
    ("code", "失物编号"),
    ("name", "物品名称"),
    ("category", "类别"),
    ("description", "特征描述"),
    ("found_location", "捡到地点"),
    ("storage_location", "存放位置"),
    ("found_time", "捡到时间"),
    ("founder", "捡到人"),
    ("status", "状态"),
    ("created_at", "登记时间"),
    ("claimer_name", "认领人姓名"),
    ("claimer_phone", "认领人电话"),
    ("claimer_group", "人群"),
    ("claimer_gender", "性别"),
    ("feature_verified", "特征已核实"),
    ("claimed_at", "认领时间"),
    ("operator", "经办人"),
    ("photo", "物品照片"),
    ("claimer_photo", "认领人照片"),
]


def export_items_to_excel(items):
    """items: sqlite3.Row 列表，导出为 Excel 下载。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "失物登记清单"

    # 标题行
    headers = [c[1] for c in COLUMNS]
    ws.append(headers)
    # 标题样式
    title_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill("solid", fgColor="2E7D9B")
    for cell in ws[1]:
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # 数据行
    for row in items:
        row = dict(row)
        line = []
        for key, _ in COLUMNS:
            val = row.get(key, "")
            if key == "feature_verified":
                val = "是" if val else ("否" if val == 0 else "")
            line.append(val if val is not None else "")
        ws.append(line)

    # 列宽
    widths = [14, 16, 12, 30, 14, 14, 16, 10, 10, 18, 12, 14, 8, 8, 10, 18, 10, 24, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 全部单元格自动换行 + 顶端对齐
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写到内存再返回
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"失物登记清单_{__import__('datetime').date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
